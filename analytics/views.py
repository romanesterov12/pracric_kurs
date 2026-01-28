"""
Views for analytics and statistics visualization.
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Avg, Count
from django.utils import timezone
from datetime import timedelta
from tracking.models import FoodLog, ActivityLog, DailyMetrics
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json


@login_required
def analytics_view(request):
    """Main analytics dashboard with charts and statistics."""
    today = timezone.now().date()

    # Date range selection
    period = request.GET.get('period', '30')
    try:
        days = int(period)
    except ValueError:
        days = 30

    start_date = today - timedelta(days=days)

    # Weight progression data
    weight_data = DailyMetrics.objects.filter(
        user=request.user,
        date__gte=start_date,
        weight__isnull=False
    ).order_by('date').values('date', 'weight')

    # Calories data (consumed vs burned)
    calories_data = []
    for day in range(days + 1):
        date = start_date + timedelta(days=day)
        consumed = FoodLog.objects.filter(
            user=request.user,
            date=date
        ).aggregate(total=Sum('calories'))['total'] or 0

        burned = ActivityLog.objects.filter(
            user=request.user,
            date=date
        ).aggregate(total=Sum('calories_burned'))['total'] or 0

        calories_data.append({
            'date': date.isoformat(),
            'consumed': consumed,
            'burned': burned
        })

    # Activity summary
    activity_summary = ActivityLog.objects.filter(
        user=request.user,
        date__gte=start_date
    ).values('activity_type').annotate(
        total_duration=Sum('duration_minutes'),
        total_calories=Sum('calories_burned'),
        count=Count('id')
    )

    # Nutrition summary
    nutrition_summary = FoodLog.objects.filter(
        user=request.user,
        date__gte=start_date
    ).aggregate(
        avg_calories=Avg('calories'),
        avg_protein=Avg('protein'),
        avg_carbs=Avg('carbs'),
        avg_fats=Avg('fats'),
        total_meals=Count('id')
    )

    # Daily metrics summary
    metrics_summary = DailyMetrics.objects.filter(
        user=request.user,
        date__gte=start_date
    ).aggregate(
        avg_sleep=Avg('sleep_hours'),
        avg_mood=Avg('mood'),
        avg_energy=Avg('energy_level'),
        avg_water=Avg('water_intake')
    )

    # Meal distribution
    meal_distribution = FoodLog.objects.filter(
        user=request.user,
        date__gte=start_date
    ).values('meal_type').annotate(
        count=Count('id'),
        total_calories=Sum('calories')
    )

    context = {
        'period': period,
        'weight_data': json.dumps(list(weight_data), default=str),
        'calories_data': json.dumps(calories_data),
        'activity_summary': activity_summary,
        'nutrition_summary': nutrition_summary,
        'metrics_summary': metrics_summary,
        'meal_distribution': meal_distribution,
    }

    return render(request, 'analytics/analytics.html', context)


@login_required
def export_xlsx_view(request):
    """Export user data to Excel file."""
    today = timezone.now().date()
    period = int(request.GET.get('period', '30'))
    start_date = today - timedelta(days=period)

    # Create workbook
    wb = Workbook()

    # Food Logs sheet
    ws_food = wb.active
    ws_food.title = "Питание"

    # Headers
    headers = ['Дата', 'Тип приема пищи', 'Продукт', 'Порция (г)', 'Калории', 'Белки', 'Углеводы', 'Жиры']
    ws_food.append(headers)

    # Style headers
    for cell in ws_food[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # Add food data
    food_logs = FoodLog.objects.filter(
        user=request.user,
        date__gte=start_date
    ).order_by('-date')

    for log in food_logs:
        ws_food.append([
            log.date.strftime('%d.%m.%Y'),
            log.get_meal_type_display(),
            log.food_name,
            log.portion_size,
            log.calories,
            log.protein,
            log.carbs,
            log.fats
        ])

    # Activity Logs sheet
    ws_activity = wb.create_sheet("Активность")

    headers = ['Дата', 'Тип активности', 'Название', 'Длительность (мин)', 'Интенсивность', 'Калории сожжено']
    ws_activity.append(headers)

    # Style headers
    for cell in ws_activity[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # Add activity data
    activity_logs = ActivityLog.objects.filter(
        user=request.user,
        date__gte=start_date
    ).order_by('-date')

    for log in activity_logs:
        ws_activity.append([
            log.date.strftime('%d.%m.%Y'),
            log.get_activity_type_display(),
            log.activity_name,
            log.duration_minutes,
            log.get_intensity_display(),
            log.calories_burned
        ])

    # Daily Metrics sheet
    ws_metrics = wb.create_sheet("Ежедневные показатели")

    headers = ['Дата', 'Вес (кг)', 'Сон (ч)', 'Настроение', 'Энергия', 'Вода (л)']
    ws_metrics.append(headers)

    # Style headers
    for cell in ws_metrics[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        cell.font = Font(bold=True, color='000000')
        cell.alignment = Alignment(horizontal='center')

    # Add metrics data
    daily_metrics = DailyMetrics.objects.filter(
        user=request.user,
        date__gte=start_date
    ).order_by('-date')

    for metric in daily_metrics:
        ws_metrics.append([
            metric.date.strftime('%d.%m.%Y'),
            metric.weight or '-',
            metric.sleep_hours or '-',
            metric.get_mood_display() if metric.mood else '-',
            metric.get_energy_level_display() if metric.energy_level else '-',
            metric.water_intake or '-'
        ])

    # Auto-adjust column widths
    for ws in [ws_food, ws_activity, ws_metrics]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=fittracker_data_{today}.xlsx'

    wb.save(response)
    return response


@login_required
def export_pdf_view(request):
    """Export user statistics to PDF file."""
    today = timezone.now().date()
    period = int(request.GET.get('period', '30'))
    start_date = today - timedelta(days=period)

    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=fittracker_report_{today}.pdf'

    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#34495E'),
        spaceAfter=12,
    )

    # Title
    story.append(Paragraph(f"Отчет FitTracker", title_style))
    story.append(
        Paragraph(f"Период: {start_date.strftime('%d.%m.%Y')} - {today.strftime('%d.%m.%Y')}", styles['Normal']))
    story.append(Paragraph(f"Пользователь: {request.user.username}", styles['Normal']))
    story.append(Spacer(1, 20))

    # Nutrition summary
    story.append(Paragraph("Сводка по питанию", heading_style))

    nutrition_data = FoodLog.objects.filter(
        user=request.user,
        date__gte=start_date
    ).aggregate(
        total_meals=Count('id'),
        avg_calories=Avg('calories'),
        avg_protein=Avg('protein'),
        avg_carbs=Avg('carbs'),
        avg_fats=Avg('fats')
    )

    nutrition_table_data = [
        ['Показатель', 'Значение'],
        ['Всего приемов пищи', str(nutrition_data['total_meals'] or 0)],
        ['Средняя калорийность',
         f"{nutrition_data['avg_calories']:.0f} ккал" if nutrition_data['avg_calories'] else '-'],
        ['Средний белок', f"{nutrition_data['avg_protein']:.1f} г" if nutrition_data['avg_protein'] else '-'],
        ['Средние углеводы', f"{nutrition_data['avg_carbs']:.1f} г" if nutrition_data['avg_carbs'] else '-'],
        ['Средние жиры', f"{nutrition_data['avg_fats']:.1f} г" if nutrition_data['avg_fats'] else '-'],
    ]

    nutrition_table = Table(nutrition_table_data, colWidths=[4 * inch, 2 * inch])
    nutrition_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    story.append(nutrition_table)
    story.append(Spacer(1, 20))

    # Activity summary
    story.append(Paragraph("Сводка по активности", heading_style))

    activity_data = ActivityLog.objects.filter(
        user=request.user,
        date__gte=start_date
    ).aggregate(
        total_workouts=Count('id'),
        total_duration=Sum('duration_minutes'),
        total_calories=Sum('calories_burned')
    )

    activity_table_data = [
        ['Показатель', 'Значение'],
        ['Всего тренировок', str(activity_data['total_workouts'] or 0)],
        ['Общая длительность', f"{activity_data['total_duration'] or 0} мин"],
        ['Всего калорий сожжено', f"{activity_data['total_calories'] or 0} ккал"],
    ]

    activity_table = Table(activity_table_data, colWidths=[4 * inch, 2 * inch])
    activity_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ECC71')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    story.append(activity_table)
    story.append(Spacer(1, 20))

    # Daily metrics summary
    story.append(Paragraph("Ежедневные показатели", heading_style))

    metrics_data = DailyMetrics.objects.filter(
        user=request.user,
        date__gte=start_date
    ).aggregate(
        avg_sleep=Avg('sleep_hours'),
        avg_mood=Avg('mood'),
        avg_energy=Avg('energy_level'),
        avg_water=Avg('water_intake')
    )

    metrics_table_data = [
        ['Показатель', 'Значение'],
        ['Средний сон', f"{metrics_data['avg_sleep']:.1f} ч" if metrics_data['avg_sleep'] else '-'],
        ['Среднее настроение', f"{metrics_data['avg_mood']:.1f}/5" if metrics_data['avg_mood'] else '-'],
        ['Средняя энергия', f"{metrics_data['avg_energy']:.1f}/5" if metrics_data['avg_energy'] else '-'],
        ['Среднее потребление воды', f"{metrics_data['avg_water']:.1f} л" if metrics_data['avg_water'] else '-'],
    ]

    metrics_table = Table(metrics_table_data, colWidths=[4 * inch, 2 * inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F39C12')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    story.append(metrics_table)

    # Build PDF
    doc.build(story)

    return response